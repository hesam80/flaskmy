import os, re, shutil , requests , sendsms
import pandas as pd
from pandas import  read_excel
import numpy as np
from bs4 import BeautifulSoup
import openpyxl
from telnetlib import Telnet


def display_menu():
	"""
		DISPLAY A MENU OF WHAT USER CAN DO
	"""
	print("************************************************")
	print("*****  Hello - Hessam Hosseini  ******")
	print("************************************************")
	selected = input(f"""WHICH ONE? (1-8)
	1) RUN pandas_practce
	2) RUN op_one
	3) RUN task to do
	4) RUN tahlil_defect
	5) RUN tahlil_defect_op1
	6) RUN str practice
	7) RUN weibull_predict
	8) Run new_analyze
	9) Run sendsms BTC pass
	PLEASE SPECIFIE WITH A NUMBER: 
	""")
	print("select Number is:",selected)
	if selected=="1":
		pandas_practce()
	elif selected=="2":
		#insert_to_excel()
		print(op_one())
	elif selected=="3":
		pandas_practce()
		task_to_do()
	elif selected=="4":
		tahlil_defect_op1()
	elif selected=="5":
		tahlil_defect()
	elif selected=="6":
		str_practice()
	elif selected=="7":
		weibull_predict()
	elif selected=="8":
		analysis()


def op_one():
	df=pd.read_excel('MON.xlsx')
	#return df['Chapter'].std()
	#tst1=df1.head()
	#print("sort date",date_sort)
	tst=df.head()
    #dp=df.groupby('Chapter').size()
	print(tst)
	#print(df.dtypes)
	#print(df.count())
def pandas_practce():
	date=tahlil_defect_op1()
	print("sort",date[0])
	#path = 'op1.xlsx'
	#df = pd.DataFrame([[date[0], date[1], date[2]]],
    #          columns=['sn off','pn off','sn on'])
	

	#df=pd.DataFrame([[date[0], date[1], date[2]]],
            #  columns=['sn off','pn off','sn on'])
	#df.to_excel('H:\op1.xlsx',sheet_name='Sheet1')

	# with pd.ExcelWriter('H:\op1.xlsx') as writer:
	#  	writer.book = openpyxl.load_workbook('H:\op1.xlsx')
	#  	df.to_excel(writer, sheet_name='Sheet2')
	# 
	#ws_dict = pd.read_excel('comptst.xlsx',Sheetname='Sheet1')
	#print(ws_dict['pn'])


	with pd.ExcelWriter('tst2.xlsx',engine='xlsxwriter') as writer:
		writer.book = openpyxl.load_workbook('tst2.xlsx')
		df.to_excel(writer)


def insert_to_excel():
	date=tahlil_defect_op1()
	df1=pd.DataFrame(date)
	#print(df1)
	df=pd.DataFrame([date[0]])
	print(df[0])
	# book=openpyxl.load_workbook('op1.xlsx')
	# writer=pd.ExcelWriter('op1.xlsx', engine='openpyxl')
	# writer.book=book
	# writer.sheets={ws.title:ws for ws in book.worksheets}
	# print("maxrow is: ",writer.sheets['Sheet1'].max_row)
	#df.to_excel(writer, sheet_name='Sheet1', index=False)
	#append_df_to_excel('op1.xlsx', df,sheet_name='Sheet1', index=False, encoding='utf-8')
	


def tahlil_defect_op1():
	df=pd.read_excel('MON.xlsx' , sheet_name='Export from MACS')
	for i in range(1):
		action=df['Action'][i]
		print(action)
		print(action.find('P/N OFF:'))
		if action.find('P/N OFF:') != '-1':
			filter_tag=action[(int(action.find('P/N OFF:'))):]
			#filter_tag.find('P/N OFF:')
			filter_1st= filter_tag.replace('P/N OFF:', '')
			filter_2nd= filter_1st.replace('S/N OFF:', '')
			filter_3nd= filter_2nd.replace('P/N ON:', '')
			filter_4nd= filter_3nd.replace('S/N ON:', '')
			expose= filter_4nd[int(filter_4nd.find('P/N'))+9:].strip()
			x =expose.split()
			i+=1
			x=pd.DataFrame(x)
			print(x)
			x.to_excel("MON.xlsx",sheet_name='Sheet_name_1')
		elif action.find('P/N OFF:') == '-1':
			#print("no part is changed")
			#print(x,len(expose))
			pass
	

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
	
	

def tahlil_defect():
	
	
	tag="""AFTER T/S FOUND DUC TEMP SENSOR FAULTY , SO REPLACED WITH S/P IAW AMM 21-60-21 P 201 CHECK FOUND OK.
P/N Name:
P/N OFF:4372C000
S/N OFF:244185-3
P/N ON:4372C000
S/N ON:1195

."""
	multi = """PACK VLV #2 REPLACED WITH S/P IAW 21-10-11 PB 401 
P/N Name:
P/N OFF:3502B000-004
S/N OFF:1074
P/N ON:3502B000-003
S/N ON:16755-3
."""
	tease="""P/N Name:
P/N OFF:aha1489
S/N OFF:2712005
P/N ON:AHA1489
S/N ON:19956."""
    
	actions = [tag,multi,tease]
	for action in actions :
		print("ok")
		print("Toole string ",action, len(actions))

	#length=len(tag)
	#print(tag+str(length))
	print("Toole string tease: ",len(tease))
	print("tease:",actions[1])
	filter_tag=tease[(-1)*(int(tag.find('P/N OFF:'))):]
	#filter_tag.find('P/N OFF:')
	filter_1st= filter_tag.replace('P/N OFF', '')
	filter_2nd= filter_1st.replace('S/N OFF', '')
	filter_3nd= filter_2nd.replace('P/N ON', '')
	filter_4nd= filter_3nd.replace('S/N ON', '')

	#print(filter_tag.replace('P/N ON:', ''))
	print("Replace string is:",filter_tag)
	print("toole string",len(multi))
	pnnumber=(-1)*(int(filter_4nd.find('P/N')-8))
	#pnnumber+=8
	print(filter_4nd[int(filter_4nd.find('P/N'))+9:].strip())
	print(len((filter_tag)))
	return filter_4nd[int(filter_4nd.find('P/N'))+9:].strip()

def str_practice():
    tease = tahlil_defect
    print(tease())

def pandasopenai():
    #from pandasai import SmartDataframe
    df = pd.DataFrame({
    "country": [
        "United States", "United Kingdom", "France", "Germany", "Italy", "Spain", "Canada", "Australia", "Japan", "China"],
    "gdp": [
        19294482071552, 2891615567872, 2411255037952, 3435817336832, 1745433788416, 1181205135360, 1607402389504, 1490967855104, 4380756541440, 14631844184064
    ],
    })
    # moghayse mafhoom jomlaat
    tags="AFTER T/S FOUND DUC TEMP SENSOR FAULTY , SO REPLACED WITH S/P IAW AMM 21-60-21 P 201 CHECK FOUND OK."

    multi = "pip install aspose-words"
    tasks={"defect1": tags, "defec2": multi}

    
    actions = ["AFTER T/S FOUND DUC TEMP SENSOR FAULTY , SO REPLACED WITH S/P IAW AMM 21-60-21 P 201 CHECK FOUND OK.","PACK VLV #2 REPLACED WITH S/P IAW 21-10-11 PB 401."]
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
    tfid=TfidfVectorizer()
    tfid_matrix=tfid.fit_transform(tasks)
    similariy=cosine_similarity(tfid_matrix[0],tfid_matrix[1])
    print(similariy)
    tfidf = TfidfVectorizer(analyzer='char_wb', ngram_range=(1, 2))
    text = ['sachin is a good player .', 'he is a good man .']
    X = tfidf.fit_transform(actions)
    print(tfidf.get_feature_names_out())

def weibull_predict():
    from predictr import Analysis
    # Data from testing
    # Failures and suspensions are lists containing the values
    
    suspensions = [1.62876357, 1.62876357, 1.62876357, 1.62876357]
    # Weibull Analysis
    x = Analysis(df=failures, ds=suspensions, show=True)
    
    x.mrr()
    x.mle()
    df = pd.DataFrame({
    "TTF": [ 
        180, 450, 740, 1040, 1360, 1700, 2070, 2460, 2890, 3380, 3870, 4430, 5070 ,5790 ,6630, 7530, 8880, 10520, 12920, 17540

    ]
    })
    print(df.head)
    #print(f"MRR: beta={x.beta:2f}, eta={x.eta:2f}\nMLE: beta={y.beta:2f}, eta={y.eta:2f}\n")
    
def new_analyze():
        # code
    import pandas as pd
    import matplotlib.pyplot as plt
    import scipy.stats as ss
    import numpy as np
    import seaborn as sns
    import weibull

    # generate standard weibull distribution of different Shape parameter
    gamma_1 = np.random.weibull(a=1,size=1000)
    gamma_half = np.random.weibull(a=0.5,size=1000)
    gamma_5 = np.random.weibull(a=5,size=1000)
    gamma_10 = np.random.weibull(a=10,size=1000)

    # plot different Weibull distribution
    sns.set_style('darkgrid')
    fig, ax = plt.subplots(2,2)
    sns.histplot(gamma_1,kde=True,ax= ax[0,0] )
    ax[0,0].set_title('Gamma = 1 ')
    sns.histplot(gamma_half,kde=True, ax= ax[0,1], legend='Y=0.5')
    ax[0,1].set_ylim([0,200])
    ax[0,1].set_title('Gamma = 0.5 ')
    sns.histplot(gamma_5,kde=True, ax= ax[1,0], legend='Y=5')
    ax[1,0].set_title('Gamma = 5 ')
    sns.histplot(gamma_10,kde=True, ax= ax[1,1], legend='Y=10')
    ax[1,1].set_title('Gamma = 10 ')
    plt.show()

    # load dataset
    specimen_strength = pd.DataFrame({
    
    "index": [
        1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20
    ],
    "TTF": [ 
        180, 450, 740, 1040, 1360, 1700, 2070, 2460, 2890, 3380, 3870, 4430, 5070 ,5790 ,6630, 7530, 8880, 10520, 12920, 17540

    ]
    
    })
    #specimen_strength = pd.read_csv('weibull.csv', header=None)
    specimen_strength.head()

    # perform weibull analysis
    analysis=weibull.Analysis(specimen_strength[0])

    # Here, we can fit using two method, mle (maximum likelihood)
    # and lr (linear regression). Generally mle is better fit
    analysis.fit(method='lr')

    # print shape parameter (Beta) and scale parameter (eta)
    print(f'shape Parameter: {analysis.beta: .02f}')
    print(f'Scale Parameter: {analysis.eta: .02f}')

    # print values of different parameters confidence interval
    analysis.stats

    # generate Weibull probplot
    analysis.probplot()

def analysis():
    from reliability.Probability_plotting import plotting_positions
    from reliability.Distributions import Weibull_Distribution
    from reliability.Probability_plotting import plot_points
    import matplotlib.pyplot as plt
    failures = [150,560,800,1720,5230,6890]
    right_censored = [340,1130,2470,4210]
    x,y=plotting_positions(failures=failures,right_censored=right_censored)

    print('x =',x)
    print('y =',y)
    dist = Weibull_Distribution(alpha=100,beta=2)
    data = dist.random_samples(1000,seed=1)

    functions = ['PDF','CDF','SF','HF','CHF']
    i = 0
    for function in functions:
        plt.subplot(151+i)
        if function == 'PDF':
            dist.PDF()
        elif function == 'CDF':
            dist.CDF()
        elif function == 'SF':
            dist.SF()
        elif function == 'HF':
            dist.HF()
        elif function == 'CHF':
            dist.CHF()
        plot_points(failures=data,func=function)
        plt.title(function)
        i+=1
    plt.gcf().set_size_inches(12,4)
    plt.tight_layout()
    plt.show()

   



#task_to_do()
#op_one()
#pandas_practce(op_one)
display_menu()
