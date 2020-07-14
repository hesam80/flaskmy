import os, re, shutil , requests , sendsms , time
import pandas as pd
from pandas import  read_excel
import numpy as np
from bs4 import BeautifulSoup
from selenium import webdriver
import openpyxl


def display_menu():
	"""
		DISPLAY A MENU OF WHAT USER CAN DO
	"""
	print("************************************************")
	print("*****  Hello - Hessam Hosseini  ******")
	print("************************************************")
	selected = input(f"""WHICH ONE? (1-8)
	1) RUN pandas_practce
	2) RUN insert into excel
	3) RUN task to do
	4) RUN tahlil_defect_op1
	5) RUN tahlil_defect
	6) RUN whatsup
	7) RUN myscrabing
	8) Run APIs
	9) Run sendsms BTC pass
	PLEASE SPECIFIE WITH A NUMBER: 
	""")
	print("select Number is:",selected)
	if selected=="1":
		pandas_practce()
	elif selected=="2":
		insert_to_excel()
		#print(op_one())
	elif selected=="3":
		pandas_practce()
		task_to_do()
	elif selected=="4":
		tahlil_defect_op1()
	elif selected=="5":
		tahlil_defect()
	elif selected=="6":
		whatsup()
	elif selected=="7":
		pass
	elif selected=="8":
		pass


def op_one():
	df=read_excel('op1.xlsx',0)
	return df['Chapter'].std()
	#tst1=df1.head()
	#print("sort date",date_sort)
	#tst=df.head()
	dp=df.groupby('Chapter').size()
	print(dp)
	#print(df.dtypes)
	#print(df.count())
def pandas_practce():
	date=tahlil_defect_op1()
	print("sort",date[0])
	#path = 'op1.xlsx'
	#df = pd.DataFrame([[date[0], date[1], date[2]]],
    #          columns=['sn off','pn off','sn on'])
	

	#df=pd.DataFrame([[date[0], date[1], date[2]]],
            #  columns=['sn off','pn off','sn on'])
	#df.to_excel('H:\op1.xlsx',sheet_name='Sheet1')

	# with pd.ExcelWriter('H:\op1.xlsx') as writer:
	#  	writer.book = openpyxl.load_workbook('H:\op1.xlsx')
	#  	df.to_excel(writer, sheet_name='Sheet2')
	# 
	#ws_dict = pd.read_excel('comptst.xlsx',Sheetname='Sheet1')
	#print(ws_dict['pn'])


	with pd.ExcelWriter('tst2.xlsx',engine='xlsxwriter') as writer:
		writer.book = openpyxl.load_workbook('tst2.xlsx')
		df.to_excel(writer)


def insert_to_excel():
	date=tahlil_defect_op1()
	df1=pd.DataFrame(date)
	#print(df1)
	df=pd.DataFrame([date[0]])
	print(df[0])
	# book=openpyxl.load_workbook('op1.xlsx')
	# writer=pd.ExcelWriter('op1.xlsx', engine='openpyxl')
	# writer.book=book
	# writer.sheets={ws.title:ws for ws in book.worksheets}
	# print("maxrow is: ",writer.sheets['Sheet1'].max_row)
	#df.to_excel(writer, sheet_name='Sheet1', index=False)
	#append_df_to_excel('op1.xlsx', df,sheet_name='Sheet1', index=False, encoding='utf-8')
	


def tahlil_defect_op1():
	df=read_excel('op1.xlsx',0)
	for i in range(6,13):
		action=df['Action'][i]
		print(action)
		print(action.find('P/N OFF:'))
		if action.find('P/N OFF:') != '-1':
			filter_tag=action[(int(action.find('P/N OFF:'))):]
			#filter_tag.find('P/N OFF:')
			filter_1st= filter_tag.replace('P/N OFF:', '')
			filter_2nd= filter_1st.replace('S/N OFF:', '')
			filter_3nd= filter_2nd.replace('P/N ON:', '')
			filter_4nd= filter_3nd.replace('S/N ON:', '')
			expose= filter_4nd[int(filter_4nd.find('P/N'))+9:].strip()
			x =expose.split()
			i+=1
			x=pd.DataFrame(x)
			print(x)
			return x
		elif action.find('P/N OFF:') == '-1':
			#print("no part is changed")
			#print(x,len(expose))
			pass
	

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
	
	

def tahlil_defect():
	
	
	tag="""TAIL NAV LIT ASSY REPLACED WITH S.B PART IAW AMM 33-42-11 CHECK FOUND OK 
P/N Name:
P/N OFF:2LA001625-07
S/N OFF:NIL
P/N ON:2LA001625-07
S/N ON:1228254
."""
	multi = """PACK VLV #2 REPLACED WITH S/P IAW 21-10-11 PB 401 
P/N Name:
P/N OFF:3502B000-004
S/N OFF:1074
P/N ON:3502B000-003
S/N ON:16755-3
."""
	tease="""P/N Name:
P/N OFF:aha1489
S/N OFF:2712005
P/N ON:AHA1489
S/N ON:19956."""

	temp="""P/N Name:
P/N OFF:
S/N OFF:
P/N ON:
S/N ON:."""


		
	#length=len(tag)
	#print(tag+str(length))
	print("toole temp",len(tag))
	print("tag:",tag)
	filter_tag=tag[(-1)*(int(tag.find('P/N OFF:'))):]
	#filter_tag.find('P/N OFF:')
	filter_1st= filter_tag.replace('P/N OFF', '')
	filter_2nd= filter_1st.replace('S/N OFF', '')
	filter_3nd= filter_2nd.replace('P/N ON', '')
	filter_4nd= filter_3nd.replace('S/N ON', '')

	#print(filter_tag.replace('P/N ON:', ''))
	print("Replace string is:",filter_tag)
	print("toole string",len(tag))
	pnnumber=(-1)*(int(filter_4nd.find('P/N')-8))
	#pnnumber+=8
	print(filter_4nd[int(filter_4nd.find('P/N'))+9:].strip())
	print(len((filter_tag)))
	return filter_4nd[int(filter_4nd.find('P/N'))+9:].strip()

def whatsup():
	from selenium import webdriver
	from selenium.webdriver.support.ui import WebDriverWait
	from selenium.webdriver.support.ui import Select
	from selenium.webdriver.common.by import By
	from selenium.webdriver.common.keys import Keys
	from selenium.webdriver.chrome.options import Options
	import time
	driver=webdriver.Firefox()
	driver.get("https://web.whatsapp.com/")
	text = "Hey, this message was sent using Selenium"
	contact="Baba"
    
	inp_xpath_search = "//input[@title='Search or start new chat']"
	input_box_search = WebDriverWait(driver,50).until(lambda driver: driver.find_element_by_xpath(inp_xpath_search))
	input_box_search.click()
	time.sleep(2)
	input_box_search.send_keys(contact)
	time.sleep(2)
	selected_contact = driver.find_element_by_xpath("//span[@title='"+contact+"']")
	selected_contact.click()
	inp_xpath = '//div[@class="_2S1VP copyable-text selectable-text"][@contenteditable="true"][@data-tab="1"]'
	input_box = driver.find_element_by_xpath(inp_xpath)
	time.sleep(2)
	input_box.send_keys(text + Keys.ENTER)
	time.sleep(2)
	driver.quit()

#task_to_do()
#op_one()
#pandas_practce(op_one)
display_menu()